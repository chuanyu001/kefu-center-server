#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
运营平台数据查询（按 VIN）
  数据源: 桌面「运营平台数据.xlsx」(按「VIN号」匹配，返回全部 19 个字段)
  支持: 命令行传 VIN / 批量从文件读 VIN / 交互输入
  输出: 终端打印 + 追加写桌面「运营平台查询结果.csv」(全部字段 + 查询时间)

用法:
  python3 运营平台查询.py LFNAHUPMXT1E19383 LFWSRX9L8TAA23391     # 直接传(可多个)
  python3 运营平台查询.py --file vins.txt                          # 批量:从文件读(txt/csv/xlsx)
  python3 运营平台查询.py                                           # 运行后按提示输入

文件输入格式:
  txt/csv : 每行一个 VIN，或空格/逗号分隔均可
  xlsx    : 取第一列的 VIN(自动跳过表头)

只依赖 openpyxl(读 xlsx)；Mac 若无,运行: pip3 install openpyxl
"""

import sys
import os
import csv
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，请先运行:  pip3 install openpyxl")
    sys.exit(1)

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
SRC = os.path.join(DESKTOP, "运营平台数据.xlsx")
OUT = os.path.join(DESKTOP, "运营平台查询结果.csv")


def load_source():
    """读 Excel → (表头list, VIN列下标, {VIN大写: 整行list})。"""
    if not os.path.exists(SRC):
        print("未找到数据源: %s" % SRC)
        sys.exit(1)
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header = [("" if h is None else str(h)) for h in next(it)]
    vin_idx = 0
    for i, h in enumerate(header):
        if "VIN" in h.upper():
            vin_idx = i
            break
    index = {}
    for row in it:
        if not row or vin_idx >= len(row) or row[vin_idx] is None:
            continue
        key = str(row[vin_idx]).strip().upper()
        if key:
            index[key] = ["" if c is None else c for c in row]
    wb.close()
    return header, vin_idx, index


def read_vins_from_file(path):
    """从文件读 VIN：xlsx 取第一列；txt/csv 按空格/逗号/换行拆。"""
    ext = os.path.splitext(path)[1].lower()
    raw = []
    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None and str(row[0]).strip():
                raw.append(str(row[0]).strip())
        wb.close()
    else:
        with open(path, encoding="utf-8-sig", errors="ignore") as f:
            text = f.read()
        for ch in [",", "，", "\n", "\t", ";", "；"]:
            text = text.replace(ch, " ")
        raw = [t for t in text.split() if t.strip()]
    out, seen = [], set()
    for v in raw:
        u = v.strip().upper()
        if not u or u in seen or u in ("VIN", "VIN号", "VINNO", "VIN码"):
            continue
        seen.add(u)
        out.append(u)
    return out


def parse_vins(text):
    for ch in [",", "，", "\n", "\t", ";", "；"]:
        text = text.replace(ch, " ")
    out, seen = [], set()
    for v in text.split():
        u = v.strip().upper()
        if u and u not in seen:
            seen.add(u)
            out.append(u)
    return out


def main():
    args = sys.argv[1:]
    vins = []
    if args and args[0] in ("--file", "-f"):
        if len(args) < 2:
            print("用法: python3 运营平台查询.py --file <文件路径>")
            return
        path = args[1]
        if not os.path.exists(path):
            print("文件不存在: %s" % path)
            return
        vins = read_vins_from_file(path)
        print("从文件读到 %d 个 VIN: %s" % (len(vins), path))
    elif args:
        vins = parse_vins(" ".join(args))
    else:
        vins = parse_vins(input("请输入 VIN 码(多个用空格/逗号分隔，或用 --file 批量): ").strip())

    if not vins:
        print("未输入 VIN，已退出。")
        return

    print("正在加载数据源(约 5 万行，稍候)…")
    header, vin_idx, index = load_source()
    print("数据源就绪，共 %d 条；开始查询 %d 个 VIN\n" % (len(index), len(vins)))

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out_header = header + ["查询时间"]
    rows = []
    hit = 0
    for v in vins:
        row = index.get(v)
        if row:
            hit += 1
            full = (row + [""] * len(header))[:len(header)]
            print("✓ %s" % v)
            for h, val in zip(header, full):
                print("    %-12s %s" % (h + ":", val if val != "" else "(空)"))
            print("")
            rows.append(full + [now])
        else:
            print("✗ %s  → 未查到\n" % v)
            miss = [""] * len(header)
            if vin_idx < len(miss):
                miss[vin_idx] = v
            rows.append(miss + ["未查到 " + now])

    new_file = not os.path.exists(OUT)
    with open(OUT, "a", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(out_header)
        w.writerows(rows)

    print("命中 %d / %d，已保存到: %s" % (hit, len(vins), OUT))


if __name__ == "__main__":
    main()
